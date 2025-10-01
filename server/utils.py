import os
from openpyxl import Workbook, load_workbook
from datetime import datetime


def ensure_workbook(path):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = 'messages'
        ws.append(['timestamp', 'type', 'number', 'body'])
        wb.save(path)


def append_message(path, mtype, number, body):
    # ensure workbook exists
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb.active
    ts = datetime.utcnow().isoformat() + 'Z'
    ws.append([ts, mtype, number, body])
    wb.save(path)
